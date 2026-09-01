#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: install openpyxl with `pip install openpyxl`") from exc


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def get_cell(row, index, ws):
    if index < 1 or row < 1:
        return ""
    if row > ws.max_row or index > ws.max_column:
        return ""
    return clean_text(ws.cell(row=row, column=index).value)


def normalise_price(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace(".", "").replace("€", "").replace("€", "").strip()
    text = text.replace(" ", "")
    text = text.replace(",", ".")
    try:
        num = float(text)
        return f"{num:.2f}".replace(".", ",") + " €"
    except ValueError:
        return text + " €" if text and not text.endswith("€") else text


def find_xlsx(path_hint=None):
    candidates = []
    if path_hint:
        candidates.append(Path(path_hint))
    candidates.extend([
        Path("plantilla-menu-teckels.xlsx"),
        Path("/Users/leandrodelmar/Documents/Codex/2026-08-19/ne/outputs/plantilla-menu-teckels.xlsx"),
        Path("/Users/leandrodelmar/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/plantilla-menu-teckels.xlsx"),
        Path("~/Desktop/plantilla-menu-teckels.xlsx").expanduser(),
        Path("~/Documents/plantilla-menu-teckels.xlsx").expanduser(),
        Path(".").resolve() / "plantilla-menu-teckels.xlsx",
    ])
    seen = set()
    for path in candidates:
        if path is None:
            continue
        try:
            resolved = path.expanduser().resolve()
        except Exception:
            resolved = path
        if str(resolved) not in seen and resolved.exists():
            seen.add(str(resolved))
            return str(resolved)
    return None


def build_menu(xlsx_path: str):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = workbook["Menu"]

    rows = []
    for row in range(5, ws.max_row + 1):
        category_id = get_cell(row, 1, ws)
        if not category_id:
            continue
        rows.append({
            "category_id": category_id,
            "category_icon": get_cell(row, 2, ws),
            "category_order": get_cell(row, 3, ws),
            "category_en": get_cell(row, 4, ws),
            "category_de": get_cell(row, 5, ws),
            "category_es": get_cell(row, 6, ws),
            "category_desc_en": get_cell(row, 7, ws),
            "category_desc_de": get_cell(row, 8, ws),
            "category_desc_es": get_cell(row, 9, ws),
            "item_order": get_cell(row, 10, ws),
            "item_en": get_cell(row, 11, ws),
            "item_de": get_cell(row, 12, ws),
            "item_es": get_cell(row, 13, ws),
            "description_en": get_cell(row, 14, ws),
            "description_de": get_cell(row, 15, ws),
            "description_es": get_cell(row, 16, ws),
            "price": get_cell(row, 17, ws),
            "visible": get_cell(row, 18, ws),
        })

    menu = []
    categories = {}
    for row in rows:
        if row["visible"].lower() != "yes":
            continue

        cat_id = row["category_id"]
        if cat_id not in categories:
            categories[cat_id] = {
                "id": cat_id,
                "e": row["category_icon"],
                "name": {"en": row["category_en"], "de": row["category_de"], "es": row["category_es"]},
                "sub": {
                    "en": row["category_desc_en"],
                    "de": row["category_desc_de"],
                    "es": row["category_desc_es"],
                },
                "items": [],
            }
            menu.append(categories[cat_id])

        item = {
            "n": {"en": row["item_en"], "de": row["item_de"], "es": row["item_es"]},
            "d": {"en": row["description_en"], "de": row["description_de"], "es": row["description_es"]},
            "p": normalise_price(row["price"]),
        }
        categories[cat_id]["items"].append(item)

    for category in menu:
        category["items"].sort(key=lambda x: x["n"]["en"])

    return menu


def write_js(menu, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(menu, ensure_ascii=False, indent=2)
    output_path.write_text(f"window.TECKELS_MENU = {payload};\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Build TECKELS menu JS data from Excel")
    parser.add_argument("xlsx_path", nargs="?", default=None, help="Path to plantilla-menu-teckels.xlsx")
    parser.add_argument("--output", default="menu-data.js", help="Output JS file")
    args = parser.parse_args()

    xlsx_path = find_xlsx(args.xlsx_path)
    if not xlsx_path:
        raise SystemExit("Could not find plantilla-menu-teckels.xlsx in the project or common folders.")

    menu = build_menu(xlsx_path)
    out = Path(args.output)
    write_js(menu, out)
    total_items = sum(len(c["items"]) for c in menu)
    print(f"Menu generated successfully: {out}, Categories: {len(menu)}, Products: {total_items}")


if __name__ == "__main__":
    main()
